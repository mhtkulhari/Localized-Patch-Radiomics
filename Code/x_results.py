from __future__ import annotations

import argparse
import re
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from all_config import (
    OUTPUT_ROOT,
    RANDOM_SEEDS,
    RANDOM_STATE,
    external_results_dir,
    results_dir,
)


METRICS = ("auc", "ap", "balanced_acc", "f1", "sensitivity")
METRIC_LABELS = {
    "auc": "AUC",
    "ap": "AP",
    "balanced_acc": "BAcc",
    "f1": "F1",
    "sensitivity": "Sensitivity",
}
INTERNAL_MODELS = (
    {
        "label": "Localized Patch (Ours)",
        "exp": "A",
        "source": "final_stacking",
        "tasks": {"cs": "cs", "zone": "zone"},
    },
    {
        "label": "Whole Organ (A)",
        "exp": "B",
        "source": "final_stacking",
        "tasks": {"cs": "cs", "zone": "zone"},
    },
    {
        "label": "Whole Organ (B)",
        "exp": "B0",
        "source": "final_stacking",
        "tasks": {"cs": "cs", "zone": "zone"},
    },
    {
        "label": "Early Feature Fusion",
        "exp": "C",
        "source": "final_stacking",
        "tasks": {"cs": "cs", "zone": "zone"},
    },
    {
        "label": "Late Prediction Fusion",
        "exp": "D",
        "source": "late_fusion",
        "tasks": {"cs": "cs", "zone": "zone"},
    },
)

TASK_LABELS = {
    "cs": "csPCa",
    "zone": "Zone",
}


def _split_mean_std(value: object) -> tuple[float, float]:
    if pd.isna(value):
        return float("nan"), float("nan")
    if isinstance(value, (int, float)):
        return float(value), 0.0
    text = str(value).strip().replace("+/-", "±")
    parts = re.split(r"\s*±\s*", text)
    if len(parts) == 2:
        return float(parts[0]), float(parts[1])
    return float(text), 0.0


def _fmt(value: object, std: float | None = None) -> str:
    if std is None:
        mean, std = _split_mean_std(value)
    else:
        mean = float(value)
    if pd.isna(mean):
        return ""
    if pd.isna(std):
        std = 0.0
    return f"{mean:.3f}±{float(std):.3f}"


def _read_mean_outer(exp: str, seed: int) -> pd.DataFrame:
    path = results_dir(exp, seed) / "final_stacking.xlsx"
    if not path.exists():
        return pd.DataFrame()
    df = pd.read_excel(path, sheet_name="mean_outer")
    if "bal_acc" in df.columns and "balanced_acc" not in df.columns:
        df = df.rename(columns={"bal_acc": "balanced_acc"})
    return df


def _read_maxpool_zone(seed: int) -> pd.DataFrame:
    path = results_dir("A", seed) / "maxpool_zone.xlsx"
    if not path.exists():
        return pd.DataFrame()
    df = pd.read_excel(path, sheet_name="mean_outer")
    if "task" not in df.columns:
        df["task"] = "zone"
    return df


def _late_fusion_internal(seed: int, task: str) -> dict[str, str] | None:
    df = _read_mean_outer("D", seed)
    if df.empty:
        return None
    task_df = df[df["task"].astype(str).str.lower().eq(task)]
    if task_df.empty:
        return None
    row = task_df.iloc[0]
    return {metric: _fmt(row.get(metric)) for metric in METRICS}


def build_internal(seed: int) -> pd.DataFrame:
    rows: list[dict[str, str]] = []
    cache: dict[tuple[str, str], pd.DataFrame] = {}

    for task in ("cs", "zone"):
        for model in INTERNAL_MODELS:
            exp = model["exp"]
            source = model["source"]
            row: dict[str, str] = {"Task": TASK_LABELS[task], "Model": model["label"]}

            if source == "late_fusion":
                metrics = _late_fusion_internal(seed, task)
                if metrics is None:
                    metrics = {metric: "" for metric in METRICS}
            else:
                key = (exp, source)
                if key not in cache:
                    cache[key] = _read_mean_outer(exp, seed)
                df = cache[key]
                if exp == "A" and task == "zone":
                    df = _read_maxpool_zone(seed)
                task_df = df[df["task"].astype(str).str.lower().eq(task)] if not df.empty else pd.DataFrame()
                if task_df.empty:
                    metrics = {metric: "" for metric in METRICS}
                else:
                    src = task_df.iloc[0]
                    metrics = {metric: _fmt(src.get(metric)) for metric in METRICS}

            row.update({METRIC_LABELS[metric]: metrics[metric] for metric in METRICS})
            rows.append(row)

    return pd.DataFrame(rows, columns=["Task", "Model", *[METRIC_LABELS[m] for m in METRICS]])


def _external_sheet(path: Path) -> tuple[str, pd.DataFrame] | None:
    if not path.exists():
        return None
    xl = pd.ExcelFile(path)
    if "late_fusion_metrics" in xl.sheet_names:
        return "late_fusion_metrics", pd.read_excel(path, sheet_name="late_fusion_metrics")
    if "metrics" in xl.sheet_names:
        return "metrics", pd.read_excel(path, sheet_name="metrics")
    return None


def _external_rows_for_seed(exp: str, seed: int) -> pd.DataFrame:
    path = external_results_dir(exp) / f"ext_results_seed{seed}.xlsx"
    loaded = _external_sheet(path)
    if loaded is None:
        return pd.DataFrame()
    _, df = loaded
    if "balanced_acc" not in df.columns and "bal_acc" in df.columns:
        df = df.rename(columns={"bal_acc": "balanced_acc"})
    df = df[df["task"].astype(str).str.lower().isin(["cs", "zone"])].copy()
    df["seed"] = seed
    return df


def _available_external_seeds() -> list[int]:
    seeds: list[int] = []
    for exp in {model["exp"] for model in INTERNAL_MODELS}:
        for path in sorted(external_results_dir(exp).glob("ext_results_seed*.xlsx")):
            match = re.search(r"seed(\d+)", path.stem)
            if match:
                seeds.append(int(match.group(1)))
    return sorted(set(seeds))


def build_external(seeds: list[int]) -> pd.DataFrame:
    cache: dict[str, pd.DataFrame] = {}
    rows: list[dict[str, object]] = []

    for task in ("cs", "zone"):
        for model in INTERNAL_MODELS:
            exp = model["exp"]
            if exp not in cache:
                frames = [_external_rows_for_seed(exp, seed) for seed in seeds]
                frames = [frame for frame in frames if not frame.empty]
                cache[exp] = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()

            df = cache[exp]
            task_df = (
                df[df["task"].astype(str).str.lower().eq(task)]
                if not df.empty else pd.DataFrame()
            )
            row = {"Task": TASK_LABELS[task], "Model": model["label"]}
            for metric in METRICS:
                vals = (
                    pd.to_numeric(task_df.get(metric, pd.Series(dtype=float)), errors="coerce")
                    .dropna()
                )
                if vals.empty:
                    row[METRIC_LABELS[metric]] = ""
                else:
                    row[METRIC_LABELS[metric]] = round(float(vals.mean()), 3)
            rows.append(row)

    return pd.DataFrame(rows, columns=["Task", "Model", *[METRIC_LABELS[m] for m in METRICS]])


def _style_sheet(writer: pd.ExcelWriter, sheet_name: str, merge_task: bool = False) -> None:
    ws = writer.book[sheet_name]
    header_fill = PatternFill("solid", fgColor="EAEAEA")
    thin_gray = Side(style="thin", color="C7C7C7")
    border = Border(bottom=thin_gray)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    headers = {cell.value: cell.column_letter for cell in ws[1]}
    if "Model" in headers:
        for cell in ws[headers["Model"]]:
            cell.alignment = Alignment(horizontal="left", vertical="center")
    widths = {"A": 13}
    for cell in ws[1]:
        if cell.value == "Model":
            widths[cell.column_letter] = 26
        elif cell.value != "Task":
            widths[cell.column_letter] = 14
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    if merge_task:
        for start, end in ((2, 6), (7, 11)):
            ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
            ws.cell(start, 1).alignment = Alignment(horizontal="center", vertical="center")


def write_results(internal: pd.DataFrame, external: pd.DataFrame, out_dir: Path) -> tuple[Path, Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    internal_path = out_dir / "internal.xlsx"
    external_path = out_dir / "external.xlsx"
    with pd.ExcelWriter(internal_path, engine="openpyxl") as writer:
        internal.to_excel(writer, index=False, sheet_name="internal")
        _style_sheet(writer, "internal", merge_task=True)
    with pd.ExcelWriter(external_path, engine="openpyxl") as writer:
        external.to_excel(writer, index=False, sheet_name="external")
        _style_sheet(writer, "external", merge_task=True)
    return internal_path, external_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Create paper-style BSPC internal/external result workbooks.")
    parser.add_argument("--seed", type=int, default=RANDOM_STATE, help="Internal result seed folder to summarize.")
    parser.add_argument("--all-seeds", action="store_true", help="Generate results for all seeds in RANDOM_SEEDS.")
    parser.add_argument(
        "--external-seeds",
        default="available",
        help="Comma-separated external seeds, 'current', or 'available'.",
    )
    parser.add_argument(
        "--out-dir",
        type=Path,
        default=None,
        help="Output directory. Defaults to OUTPUT_ROOT/results/seed{seed}.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    seeds = RANDOM_SEEDS if args.all_seeds else [args.seed]
    for seed in seeds:
        if args.external_seeds == "current":
            external_seeds = [seed]
        elif args.external_seeds == "available":
            external_seeds = _available_external_seeds() or [seed]
        else:
            external_seeds = [int(x.strip()) for x in args.external_seeds.split(",") if x.strip()]

        if args.out_dir is not None and args.all_seeds:
            out_dir = args.out_dir / f"seed{seed}"
        else:
            out_dir = args.out_dir or (OUTPUT_ROOT / "results" / f"seed{seed}")
        internal = build_internal(seed)
        external = build_external(external_seeds)
        internal_path, external_path = write_results(internal, external, out_dir)

        print(f"[RESULTS] wrote {internal_path}")
        print(f"[RESULTS] wrote {external_path}")
        print(f"[RESULTS] internal seed: {seed}")
        print(f"[RESULTS] external seeds: {external_seeds}")


if __name__ == "__main__":
    main()
